from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from openpyxl import load_workbook

@login_required
def index(request):
    return render(request, 'accounting_system/index.html')

@login_required
def chart_of_accounts(request):
    wb = load_workbook(filename='accounting_system/AIS2.xlsx')
    sheet = wb['Chart Of Accounts']
    rows = []

    # Manual JSON Conversion
    for cell in sheet.iter_rows(max_col=2):
        if type(cell[1]).__name__ == 'MergedCell':
            rows.append([cell[0].value])
        else:
            rows[-1].append((cell[0].value, cell[1].value))

    context = {
        'rows': rows,
    }
    return render(request, 'accounting_system/chartofaccounts.html', context)
